from openpyxl import load_workbook

import mysql.connector

mydb = mysql.connector.connect(
  host="localhost",
  user="deepak",
  password="deepak",
  database="project"
)

wb = load_workbook('MyFoodData.xlsx')

sheet = wb[wb.sheetnames[0]]

mycursor = mydb.cursor()

mycursor.execute("CREATE TABLE  IF NOT EXISTS MyFoodData (ID INT NOT NULL AUTO_INCREMENT, Name VARCHAR(200), Food_Group VARCHAR(50), Fat DECIMAL(10,2), Protein DECIMAL(10,2), Carbohydrate DECIMAL(10,2), PRIMARY KEY(ID))")

for i in range(5,14169):
	if sheet.cell(i,3).value in ['Vegetables','Grains and Pasta','Fats and Oils','Meats','Dairy and Egg Products']:
		mycursor.execute("INSERT INTO MyFoodData (Name, Food_Group, Fat, Protein, Carbohydrate) VALUES (%s,%s,%s,%s,%s)",(sheet.cell(i,2).value,sheet.cell(i,3).value,sheet.cell(i,5).value,sheet.cell(i,6).value,sheet.cell(i,7).value,))

mydb.commit()

mydb.close()